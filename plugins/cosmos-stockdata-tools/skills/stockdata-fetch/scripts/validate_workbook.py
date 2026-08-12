#!/usr/bin/env python3
"""Validate a stockdata-fetch SuperMind workbook without another data-source call."""

from __future__ import annotations

import argparse
import ast
import json
import re
from dataclasses import asdict, dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable


EXPECTED_SHEETS = (
    "market",
    "sentiment",
    "theme_index",
    "theme_members",
    "mtss",
    "stocks_meta",
    "valuation",
    "indexes_all",
    "indexes_meta",
)


@dataclass(frozen=True)
class ValidationContract:
    expected_indexes: int
    min_all_indexes: int
    min_quoted_indexes: int
    min_stocks: int = 5000
    min_industry_coverage: float = 0.95
    min_concept_coverage: float = 0.90


@dataclass(frozen=True)
class CheckResult:
    code: str
    passed: bool
    detail: str


@dataclass(frozen=True)
class ValidationResult:
    workbook: str
    expected_date: str
    checks: tuple[CheckResult, ...]

    @property
    def passed(self) -> bool:
        return all(check.passed for check in self.checks)

    def to_dict(self) -> dict[str, Any]:
        return {
            "workbook": self.workbook,
            "expected_date": self.expected_date,
            "passed": self.passed,
            "checks": [asdict(check) for check in self.checks],
        }


def _cell_source(cell: dict[str, Any]) -> str:
    source = cell.get("source", "")
    return "".join(source) if isinstance(source, list) else str(source)


def _literal_notebook_assignment(notebook_path: Path, name: str) -> Any:
    notebook = json.loads(notebook_path.read_text(encoding="utf-8"))
    code = "\n".join(
        _cell_source(cell)
        for cell in notebook.get("cells", [])
        if cell.get("cell_type") == "code"
    )
    tree = ast.parse(code)
    values = []
    for node in ast.walk(tree):
        if not isinstance(node, (ast.Assign, ast.AnnAssign)):
            continue
        targets = node.targets if isinstance(node, ast.Assign) else [node.target]
        if any(isinstance(target, ast.Name) and target.id == name for target in targets):
            values.append(ast.literal_eval(node.value))
    if len(values) != 1:
        raise ValueError(f"packaged notebook must define exactly one literal {name}")
    return values[0]


def expected_index_count(notebook_path: Path) -> int:
    """Read INDEX_CODES from the packaged notebook instead of duplicating its length."""
    value = _literal_notebook_assignment(notebook_path, "INDEX_CODES")
    if not isinstance(value, list):
        raise ValueError("packaged notebook must define exactly one literal INDEX_CODES list")
    return len(value)


def expected_all_index_minimum(notebook_path: Path) -> int:
    """Read the full index-catalog floor from the packaged notebook."""
    value = _literal_notebook_assignment(notebook_path, "MIN_ALL_INDEXES")
    if isinstance(value, bool) or not isinstance(value, int) or value <= 0:
        raise ValueError("packaged notebook MIN_ALL_INDEXES must be a positive integer")
    return value


def expected_quoted_index_minimum(notebook_path: Path) -> int:
    """Read the quoted-index floor from the packaged notebook."""
    value = _literal_notebook_assignment(notebook_path, "MIN_QUOTED_INDEXES")
    if isinstance(value, bool) or not isinstance(value, int) or value <= 0:
        raise ValueError("packaged notebook MIN_QUOTED_INDEXES must be a positive integer")
    return value


def _expected_date(path: Path, explicit: str | None) -> str:
    if explicit:
        parsed = datetime.strptime(explicit, "%Y-%m-%d").date()
        return parsed.isoformat()
    match = re.fullmatch(r"supermind_full_(\d{8})\.xlsx", path.name)
    if not match:
        raise ValueError("pass --date when the workbook name is not supermind_full_YYYYMMDD.xlsx")
    return datetime.strptime(match.group(1), "%Y%m%d").date().isoformat()


def _normalize_date(value: Any) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", text):
        raise ValueError(f"invalid market date value: {text!r}")
    return datetime.strptime(text, "%Y-%m-%d").date().isoformat()


def _rows(sheet: Any) -> tuple[dict[str, int], Iterable[tuple[Any, ...]]]:
    rows = sheet.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        return {}, ()
    indexes = {str(name): idx for idx, name in enumerate(header) if name is not None}
    return indexes, rows


def _data_row_count(sheet: Any) -> int:
    _, rows = _rows(sheet)
    return sum(1 for row in rows if any(value is not None for value in row))


def _value(row: tuple[Any, ...], indexes: dict[str, int], column: str) -> Any:
    if column not in indexes:
        raise ValueError(f"missing required column: {column}")
    index = indexes[column]
    return row[index] if index < len(row) else None


def validate_workbook(
    workbook_path: Path,
    *,
    expected_date: str | None = None,
    contract: ValidationContract | None = None,
    notebook_path: Path | None = None,
) -> ValidationResult:
    """Return every structural and notebook-equivalent validation check."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - exercised by runtime environments
        raise RuntimeError("openpyxl is required; install the bundled requirements.txt") from exc

    workbook_path = Path(workbook_path).expanduser().resolve()
    day = _expected_date(workbook_path, expected_date)
    if contract is None:
        source = notebook_path or Path(__file__).with_name("extract_daily.ipynb")
        contract = ValidationContract(
            expected_indexes=expected_index_count(source),
            min_all_indexes=expected_all_index_minimum(source),
            min_quoted_indexes=expected_quoted_index_minimum(source),
        )

    checks: list[CheckResult] = []
    try:
        book = load_workbook(workbook_path, read_only=True, data_only=True)
    except Exception as exc:
        return ValidationResult(
            str(workbook_path),
            day,
            (CheckResult("workbook_opens", False, f"{type(exc).__name__}: {exc}"),),
        )

    try:
        checks.append(CheckResult("workbook_opens", True, "opened successfully"))
        missing = sorted(set(EXPECTED_SHEETS) - set(book.sheetnames))
        checks.append(CheckResult("expected_sheets", not missing, f"missing={missing}"))
        if missing:
            return ValidationResult(str(workbook_path), day, tuple(checks))

        market = book["market"]
        indexes, rows = _rows(market)
        required = {
            "date",
            "asset_type",
            "high",
            "low",
            "industry_ths_l1",
            "concepts",
        }
        missing_columns = sorted(required - set(indexes))
        if missing_columns:
            checks.append(
                CheckResult("market_columns", False, f"missing={missing_columns}")
            )
            return ValidationResult(str(workbook_path), day, tuple(checks))

        stock_count = 0
        index_count = 0
        industry_count = 0
        concept_count = 0
        dates: set[str] = set()
        high_low_ok = True
        for row in rows:
            if not any(value is not None for value in row):
                continue
            dates.add(_normalize_date(_value(row, indexes, "date")))
            asset_type = _value(row, indexes, "asset_type")
            if asset_type == "index":
                index_count += 1
                continue
            if asset_type != "stock":
                continue
            stock_count += 1
            if _value(row, indexes, "industry_ths_l1") is not None:
                industry_count += 1
            if _value(row, indexes, "concepts") is not None:
                concept_count += 1
            high = _value(row, indexes, "high")
            low = _value(row, indexes, "low")
            if high is not None and low is not None and float(high) < float(low):
                high_low_ok = False

        industry_coverage = industry_count / stock_count if stock_count else 0.0
        concept_coverage = concept_count / stock_count if stock_count else 0.0
        checks.extend(
            [
                CheckResult(
                    "stock_rows",
                    stock_count >= contract.min_stocks,
                    f"actual={stock_count}, minimum={contract.min_stocks}",
                ),
                CheckResult(
                    "common_indexes",
                    index_count == contract.expected_indexes,
                    f"actual={index_count}, expected={contract.expected_indexes}",
                ),
                CheckResult(
                    "market_date",
                    dates == {day},
                    f"actual={sorted(dates)}, expected={[day]}",
                ),
                CheckResult("high_low", high_low_ok, "all non-null high >= low"),
                CheckResult(
                    "industry_coverage",
                    industry_coverage > contract.min_industry_coverage,
                    f"actual={industry_coverage:.3%}, minimum>{contract.min_industry_coverage:.1%}",
                ),
                CheckResult(
                    "concept_coverage",
                    concept_coverage > contract.min_concept_coverage,
                    f"actual={concept_coverage:.3%}, minimum>{contract.min_concept_coverage:.1%}",
                ),
            ]
        )

        for code, sheet_name, minimum in (
            ("theme_members", "theme_members", 1),
            ("mtss", "mtss", 1),
            ("valuation", "valuation", 1),
        ):
            actual = _data_row_count(book[sheet_name])
            checks.append(
                CheckResult(code, actual >= minimum, f"actual={actual}, minimum={minimum}")
            )

        all_indexes = book["indexes_all"]
        index_columns, index_rows = _rows(all_indexes)
        missing_index_columns = sorted({"close", "has_quote"} - set(index_columns))
        checks.append(
            CheckResult(
                "index_columns",
                not missing_index_columns,
                f"missing={missing_index_columns}",
            )
        )
        if not missing_index_columns:
            all_index_count = 0
            quoted_index_count = 0
            quote_flag_ok = True
            for row in index_rows:
                if not any(value is not None for value in row):
                    continue
                all_index_count += 1
                close = _value(row, index_columns, "close")
                has_quote = _value(row, index_columns, "has_quote")
                expected_has_quote = close is not None
                if not isinstance(has_quote, bool) or has_quote != expected_has_quote:
                    quote_flag_ok = False
                quoted_index_count += int(expected_has_quote)
            checks.extend(
                [
                    CheckResult(
                        "all_indexes",
                        all_index_count >= contract.min_all_indexes,
                        f"actual={all_index_count}, minimum={contract.min_all_indexes}",
                    ),
                    CheckResult(
                        "quoted_indexes",
                        quoted_index_count >= contract.min_quoted_indexes,
                        f"actual={quoted_index_count}, minimum={contract.min_quoted_indexes}",
                    ),
                    CheckResult(
                        "index_quote_flag",
                        quote_flag_ok,
                        "has_quote exactly matches whether close is non-null",
                    ),
                ]
            )
    except (TypeError, ValueError) as exc:
        checks.append(CheckResult("workbook_schema", False, str(exc)))
    finally:
        book.close()

    return ValidationResult(str(workbook_path), day, tuple(checks))


def print_result(result: ValidationResult) -> None:
    for check in result.checks:
        label = "PASS" if check.passed else "FAIL"
        print(f"[{label}] {check.code}: {check.detail}")
    print("VALIDATION PASSED" if result.passed else "VALIDATION FAILED")


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--date", help="Expected Beijing date in YYYY-MM-DD form")
    parser.add_argument("--json", action="store_true", help="Print machine-readable JSON")
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    result = validate_workbook(args.workbook, expected_date=args.date)
    if args.json:
        print(json.dumps(result.to_dict(), ensure_ascii=False, indent=2))
    else:
        print_result(result)
    return 0 if result.passed else 1


if __name__ == "__main__":
    raise SystemExit(main())
