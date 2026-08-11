from __future__ import annotations

import sys
import tempfile
import unittest
import json
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook


SCRIPTS = Path(__file__).resolve().parents[1] / "scripts"
sys.path.insert(0, str(SCRIPTS))

from validate_workbook import (  # noqa: E402
    ValidationContract,
    expected_all_index_minimum,
    expected_index_count,
    expected_quoted_index_minimum,
    validate_workbook,
)


CONTRACT = ValidationContract(
    min_stocks=2,
    expected_indexes=2,
    min_industry_coverage=0.5,
    min_concept_coverage=0.5,
    min_all_indexes=2,
    min_quoted_indexes=1,
)


def create_valid_workbook(path: Path) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)

    market = workbook.create_sheet("market")
    market.append(["date", "asset_type", "high", "low", "industry_ths_l1", "concepts"])
    market.append(["2026-08-07", "stock", 11.0, 9.0, "行业A", "概念A"])
    market.append(["2026-08-07", "stock", 12.0, 8.0, "行业B", "概念B"])
    market.append(["2026-08-07", "index", 1.0, 1.0, None, None])
    market.append(["2026-08-07", "index", 1.0, 1.0, None, None])

    for name in (
        "sentiment",
        "theme_index",
        "theme_members",
        "mtss",
        "stocks_meta",
        "valuation",
        "indexes_meta",
    ):
        sheet = workbook.create_sheet(name)
        sheet.append(["value"])
        sheet.append([1])
    indexes_all = workbook.create_sheet("indexes_all")
    indexes_all.append(["symbol", "close", "has_quote"])
    indexes_all.append(["index-a", 100.0, True])
    indexes_all.append(["index-b", None, False])
    workbook.save(path)


class ValidateWorkbookTests(unittest.TestCase):
    def setUp(self) -> None:
        self.tempdir = tempfile.TemporaryDirectory()
        self.path = Path(self.tempdir.name) / "supermind_full_20260807.xlsx"
        create_valid_workbook(self.path)

    def tearDown(self) -> None:
        self.tempdir.cleanup()

    def validate(self):
        return validate_workbook(self.path, contract=CONTRACT)

    def failures(self) -> set[str]:
        return {check.code for check in self.validate().checks if not check.passed}

    def edit(self, callback) -> None:
        workbook = load_workbook(self.path)
        callback(workbook)
        workbook.save(self.path)

    def test_valid_workbook_passes(self) -> None:
        self.assertTrue(self.validate().passed)

    def test_missing_sheet_fails(self) -> None:
        self.edit(lambda workbook: workbook.remove(workbook["valuation"]))
        self.assertIn("expected_sheets", self.failures())

    def test_stock_row_minimum_fails(self) -> None:
        self.edit(lambda workbook: workbook["market"].delete_rows(3))
        self.assertIn("stock_rows", self.failures())

    def test_exact_common_index_count_fails(self) -> None:
        self.edit(lambda workbook: workbook["market"].delete_rows(5))
        self.assertIn("common_indexes", self.failures())

    def test_wrong_market_date_fails(self) -> None:
        self.edit(lambda workbook: setattr(workbook["market"]["A2"], "value", "2026-08-06"))
        self.assertIn("market_date", self.failures())

    def test_high_below_low_fails(self) -> None:
        self.edit(lambda workbook: setattr(workbook["market"]["C2"], "value", 1.0))
        self.assertIn("high_low", self.failures())

    def test_industry_coverage_fails(self) -> None:
        def clear(workbook):
            workbook["market"]["E2"] = None
            workbook["market"]["E3"] = None

        self.edit(clear)
        self.assertIn("industry_coverage", self.failures())

    def test_concept_coverage_fails(self) -> None:
        def clear(workbook):
            workbook["market"]["F2"] = None
            workbook["market"]["F3"] = None

        self.edit(clear)
        self.assertIn("concept_coverage", self.failures())

    def test_empty_theme_members_fails(self) -> None:
        self.edit(lambda workbook: workbook["theme_members"].delete_rows(2))
        self.assertIn("theme_members", self.failures())

    def test_empty_mtss_fails(self) -> None:
        self.edit(lambda workbook: workbook["mtss"].delete_rows(2))
        self.assertIn("mtss", self.failures())

    def test_full_index_minimum_fails(self) -> None:
        self.edit(lambda workbook: workbook["indexes_all"].delete_rows(3))
        self.assertIn("all_indexes", self.failures())

    def test_quoted_index_minimum_fails(self) -> None:
        def remove_quotes(workbook):
            workbook["indexes_all"]["B2"] = None
            workbook["indexes_all"]["C2"] = False

        self.edit(remove_quotes)
        self.assertIn("quoted_indexes", self.failures())

    def test_index_quote_flag_must_match_close_presence(self) -> None:
        self.edit(lambda workbook: setattr(workbook["indexes_all"]["C3"], "value", True))
        self.assertIn("index_quote_flag", self.failures())

    def test_missing_index_quote_column_fails(self) -> None:
        self.edit(lambda workbook: setattr(workbook["indexes_all"]["C1"], "value", "wrong"))
        self.assertIn("index_columns", self.failures())

    def test_empty_valuation_fails(self) -> None:
        self.edit(lambda workbook: workbook["valuation"].delete_rows(2))
        self.assertIn("valuation", self.failures())

    def test_explicit_date_rejects_wrong_filename_date(self) -> None:
        result = validate_workbook(
            self.path,
            expected_date="2026-08-06",
            contract=CONTRACT,
        )
        failures = {check.code for check in result.checks if not check.passed}
        self.assertIn("market_date", failures)

    def test_corrupt_workbook_fails_closed(self) -> None:
        self.path.write_bytes(b"not an xlsx file")
        self.assertEqual(self.failures(), {"workbook_opens"})

    def test_missing_required_market_column_fails(self) -> None:
        self.edit(lambda workbook: setattr(workbook["market"]["A1"], "value", "wrong"))
        self.assertIn("market_columns", self.failures())

    def test_coverage_equal_to_threshold_fails(self) -> None:
        self.edit(lambda workbook: setattr(workbook["market"]["E2"], "value", None))
        self.assertIn("industry_coverage", self.failures())

    def test_date_string_with_suffix_fails(self) -> None:
        self.edit(
            lambda workbook: setattr(
                workbook["market"]["A2"], "value", "2026-08-07junk"
            )
        )
        self.assertIn("workbook_schema", self.failures())

    def test_timestamp_string_fails(self) -> None:
        self.edit(
            lambda workbook: setattr(
                workbook["market"]["A2"], "value", "2026-08-07 00:00:00"
            )
        )
        self.assertIn("workbook_schema", self.failures())

    def test_date_and_datetime_cells_are_normalized(self) -> None:
        def replace_dates(workbook):
            workbook["market"]["A2"] = date(2026, 8, 7)
            workbook["market"]["A3"] = datetime(2026, 8, 7, 12, 30)
            workbook["market"]["A4"] = date(2026, 8, 7)
            workbook["market"]["A5"] = datetime(2026, 8, 7, 23, 59)

        self.edit(replace_dates)
        self.assertTrue(self.validate().passed)

    def test_index_codes_must_be_one_literal_list(self) -> None:
        notebook_path = Path(self.tempdir.name) / "bad.ipynb"
        notebook_path.write_text(
            json.dumps(
                {
                    "cells": [
                        {"cell_type": "code", "source": "INDEX_CODES = tuple(['a'])"}
                    ]
                }
            ),
            encoding="utf-8",
        )
        with self.assertRaises((ValueError, TypeError)):
            expected_index_count(notebook_path)

    def test_packaged_notebook_defines_and_checks_quoted_index_minimum(self) -> None:
        notebook_path = SCRIPTS / "extract_daily.ipynb"
        self.assertEqual(expected_all_index_minimum(notebook_path), 20000)
        self.assertEqual(expected_quoted_index_minimum(notebook_path), 4000)
        notebook = json.loads(notebook_path.read_text(encoding="utf-8"))
        source = "\n".join(
            "".join(cell.get("source", ""))
            if isinstance(cell.get("source", ""), list)
            else str(cell.get("source", ""))
            for cell in notebook["cells"]
            if cell.get("cell_type") == "code"
        )
        self.assertIn("len(df_indexes_all) >= MIN_ALL_INDEXES", source)
        self.assertIn("quoted_index_count < MIN_QUOTED_INDEXES", source)
        self.assertIn("has_quote 与 close 非空状态不一致", source)


if __name__ == "__main__":
    unittest.main()
