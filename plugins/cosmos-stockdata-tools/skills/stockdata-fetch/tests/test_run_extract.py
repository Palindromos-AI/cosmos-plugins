from __future__ import annotations

import copy
import io
import json
import os
import sys
import tempfile
import unittest
from contextlib import redirect_stderr, redirect_stdout
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import patch

from openpyxl import Workbook, load_workbook


SCRIPTS = Path(__file__).resolve().parents[1] / "scripts"
sys.path.insert(0, str(SCRIPTS))

import run_extract as driver  # noqa: E402


def notebook(source):
    return {"cells": [{"cell_type": "code", "source": source, "outputs": []}]}


class FakeWebSocket:
    def __init__(self, fail_after: int | None = None, received=None):
        self.sent = []
        self.fail_after = fail_after
        self.closed = False
        self.received = list(received or [])
        self.timeout = None

    def send(self, value):
        if self.fail_after is not None and len(self.sent) >= self.fail_after:
            raise RuntimeError("send failed")
        self.sent.append(value)

    def close(self):
        self.closed = True

    def settimeout(self, value):
        self.timeout = value

    def recv(self):
        if not self.received:
            raise AssertionError("no queued websocket message")
        value = self.received.pop(0)
        if isinstance(value, BaseException):
            raise value
        return value


class RunExtractTests(unittest.TestCase):
    def setUp(self) -> None:
        self.tempdir = tempfile.TemporaryDirectory()
        driver.CONFIG = driver.RuntimeConfig(
            output_dir=Path(self.tempdir.name),
            initial_watch_seconds=0,
        )
        driver._USER_CACHE = None
        driver._TOKEN_CACHE = None

    def tearDown(self) -> None:
        self.tempdir.cleanup()

    def test_requires_exactly_one_default_target(self) -> None:
        driver.require_default_target(notebook("TARGET_DATE = None\n"))
        with self.assertRaises(driver.DriverError):
            driver.require_default_target(notebook("TARGET_DATE = '2026-08-07'\n"))

    def test_injects_date_without_mutating_canonical_notebook(self) -> None:
        canonical = notebook(["TARGET_DATE = None  # latest\n", "print(TARGET_DATE)\n"])
        original = copy.deepcopy(canonical)
        overridden = driver.notebook_for_date(canonical, "2026-08-07")
        self.assertEqual(canonical, original)
        self.assertIn("TARGET_DATE = '2026-08-07'", driver._cell_source(overridden["cells"][0]))
        self.assertIsInstance(overridden["cells"][0]["source"], list)

    def test_rejects_nonexistent_calendar_date(self) -> None:
        with self.assertRaises(driver.DriverError):
            driver.validate_date("2026-02-30")

    def test_token_prefers_environment(self) -> None:
        token_file = Path(self.tempdir.name) / "token.txt"
        token_file.write_text("file-token", encoding="utf-8")
        driver.CONFIG.token_file = token_file
        with patch.dict(os.environ, {"SUPERMIND_TOKEN": "env-token"}):
            self.assertEqual(driver.token(), "env-token")

    def test_token_reads_explicit_file(self) -> None:
        token_file = Path(self.tempdir.name) / "token.txt"
        token_file.write_text("file-token\n", encoding="utf-8")
        driver.CONFIG.token_file = token_file
        with patch.dict(os.environ, {}, clear=True):
            self.assertEqual(driver.token(), "file-token")

    def test_missing_token_fails_without_reading_plugin_files(self) -> None:
        with patch.dict(os.environ, {}, clear=True):
            with self.assertRaises(driver.DriverError):
                driver.token()

    def test_parser_defaults_to_each_users_supermind_config_file(self) -> None:
        with patch.dict(os.environ, {}, clear=True):
            args = driver.build_parser().parse_args(["status"])
        self.assertEqual(args.token_file, Path.home() / ".config" / "supermind" / "token")

    def test_token_file_environment_overrides_user_config_default(self) -> None:
        configured = Path(self.tempdir.name) / "custom-token"
        with patch.dict(
            os.environ,
            {"SUPERMIND_TOKEN_FILE": str(configured)},
            clear=True,
        ):
            args = driver.build_parser().parse_args(["status"])
        self.assertEqual(args.token_file, configured)

    def test_explicit_token_file_overrides_environment_path(self) -> None:
        configured = Path(self.tempdir.name) / "configured-token"
        explicit = Path(self.tempdir.name) / "explicit-token"
        with patch.dict(
            os.environ,
            {"SUPERMIND_TOKEN_FILE": str(configured)},
            clear=True,
        ):
            args = driver.build_parser().parse_args(
                ["--token-file", str(explicit), "status"]
            )
        self.assertEqual(args.token_file, explicit)

    def test_user_name_is_discovered_and_cached(self) -> None:
        with patch.object(driver, "hub_user", return_value={"name": "alice"}) as lookup:
            self.assertEqual(driver.user_name(), "alice")
            self.assertEqual(driver.user_name(), "alice")
            lookup.assert_called_once()

    def test_run_checks_local_dependencies_before_remote_mutation(self) -> None:
        with (
            patch.object(
                driver,
                "require_dependencies",
                side_effect=driver.DriverError("missing websocket-client"),
            ),
            patch.object(driver, "load_notebook") as load,
            patch.object(driver, "start_server") as start,
            patch.object(driver, "push_notebook") as push,
            patch.object(driver, "api") as api,
        ):
            with self.assertRaisesRegex(driver.DriverError, "websocket-client"):
                driver.run()
        load.assert_not_called()
        start.assert_not_called()
        push.assert_not_called()
        api.assert_not_called()
        self.assertFalse(driver.state_path().exists())

    def test_fetch_checks_local_dependencies_before_remote_access(self) -> None:
        with (
            patch.object(
                driver,
                "require_dependencies",
                side_effect=driver.DriverError("missing openpyxl"),
            ),
            patch.object(driver, "start_server") as start,
            patch.object(driver, "api") as api,
        ):
            with self.assertRaisesRegex(driver.DriverError, "openpyxl"):
                driver.fetch("2026-08-07", allow_existing=True)
        start.assert_not_called()
        api.assert_not_called()

    def test_dependency_preflight_rejects_import_failure(self) -> None:
        with patch.object(
            driver.importlib,
            "import_module",
            side_effect=OSError("broken installation"),
        ):
            with self.assertRaisesRegex(driver.DriverError, "import failed"):
                driver.require_dependencies("websocket")

    def test_dependency_preflight_rejects_wrong_distribution_version(self) -> None:
        websocket = SimpleNamespace(
            create_connection=object(),
            WebSocketTimeoutException=TimeoutError,
        )
        with (
            patch.object(driver.importlib, "import_module", return_value=websocket),
            patch.object(driver.importlib.metadata, "version", return_value="0.0.0"),
        ):
            with self.assertRaisesRegex(driver.DriverError, "expected 1.8.0"):
                driver.require_dependencies("websocket")

    def test_dependency_preflight_rejects_wrong_websocket_module(self) -> None:
        with (
            patch.object(driver.importlib, "import_module", return_value=SimpleNamespace()),
            patch.object(driver.importlib.metadata, "version", return_value="1.8.0"),
        ):
            with self.assertRaisesRegex(driver.DriverError, "missing API"):
                driver.require_dependencies("websocket")

    def test_watch_checks_dependencies_before_kernel_access(self) -> None:
        with (
            patch.object(
                driver,
                "require_dependencies",
                side_effect=driver.DriverError("missing websocket-client"),
            ),
            patch.object(driver, "kernels") as kernels,
        ):
            with self.assertRaisesRegex(driver.DriverError, "websocket-client"):
                driver.watch(30)
        kernels.assert_not_called()

    def test_exec_checks_dependencies_before_server_access(self) -> None:
        with (
            patch.object(
                driver,
                "require_dependencies",
                side_effect=driver.DriverError("missing websocket-client"),
            ),
            patch.object(driver, "start_server") as start,
        ):
            with self.assertRaisesRegex(driver.DriverError, "websocket-client"):
                driver.exec_code("1 + 1")
        start.assert_not_called()

    def _run_with(self, fake_ws: FakeWebSocket):
        canonical = notebook("TARGET_DATE = None\nprint('one')\n")
        pushes = []
        deletes = []

        def fake_api(method, path, data=None, **kwargs):
            if method == "POST" and path.endswith("/api/kernels"):
                return {"id": "kernel-1"}
            if method == "DELETE":
                deletes.append(path)
                return None
            raise AssertionError((method, path, data, kwargs))

        with (
            patch.object(driver, "load_notebook", return_value=canonical),
            patch.object(driver, "start_server"),
            patch.object(driver, "acquire_account_run_lock", return_value=object()),
            patch.object(driver, "release_account_run_lock"),
            patch.object(driver, "kernels", return_value=[]),
            patch.object(driver, "cloud_workbooks", return_value=[]),
            patch.object(driver, "push_notebook", side_effect=lambda value: pushes.append(copy.deepcopy(value))),
            patch.object(driver, "api", side_effect=fake_api),
            patch.object(driver, "_ws_connect", return_value=fake_ws),
            patch.object(driver, "user_name", return_value="alice"),
        ):
            driver.run("2026-08-07")
        return canonical, pushes, deletes

    def test_historical_run_restores_canonical_cloud_copy(self) -> None:
        canonical, pushes, deletes = self._run_with(FakeWebSocket())
        self.assertEqual(len(pushes), 2)
        self.assertIn("TARGET_DATE = '2026-08-07'", driver._cell_source(pushes[0]["cells"][0]))
        self.assertEqual(pushes[1], canonical)
        self.assertEqual(deletes, [])
        state = json.loads((Path(self.tempdir.name) / ".runstate.json").read_text(encoding="utf-8"))
        self.assertEqual(state["target_date"], "2026-08-07")
        self.assertEqual(state["phase"], "submitted")
        self.assertEqual(state["baseline_versions"], {})

    def test_failed_submission_restores_canonical_and_deletes_kernel(self) -> None:
        canonical = notebook("TARGET_DATE = None\nprint('one')\n")
        pushes = []
        deletes = []
        fake_ws = FakeWebSocket(fail_after=0)

        def fake_api(method, path, data=None, **kwargs):
            if method == "POST":
                return {"id": "kernel-1"}
            if method == "DELETE":
                deletes.append(path)
                return None
            raise AssertionError((method, path, data, kwargs))

        with (
            patch.object(driver, "load_notebook", return_value=canonical),
            patch.object(driver, "start_server"),
            patch.object(driver, "acquire_account_run_lock", return_value=object()),
            patch.object(driver, "release_account_run_lock"),
            patch.object(driver, "kernels", return_value=[]),
            patch.object(driver, "cloud_workbooks", return_value=[]),
            patch.object(driver, "push_notebook", side_effect=lambda value: pushes.append(copy.deepcopy(value))),
            patch.object(driver, "api", side_effect=fake_api),
            patch.object(driver, "_ws_connect", return_value=fake_ws),
            patch.object(driver, "user_name", return_value="alice"),
        ):
            with self.assertRaises(RuntimeError):
                driver.run("2026-08-07")
        self.assertEqual(pushes[-1], canonical)
        self.assertEqual(len(deletes), 1)
        state = json.loads(driver.state_path().read_text(encoding="utf-8"))
        self.assertEqual(state["phase"], "aborted")
        self.assertEqual(state["failed_phase"], "submitting")
        self.assertIn("send failed", state["error"])

    def test_failed_push_without_kernel_records_aborted_state(self) -> None:
        canonical = notebook("TARGET_DATE = None\n")
        with (
            patch.object(driver, "load_notebook", return_value=canonical),
            patch.object(driver, "start_server"),
            patch.object(driver, "acquire_account_run_lock", return_value=object()),
            patch.object(driver, "release_account_run_lock"),
            patch.object(driver, "kernels", return_value=[]),
            patch.object(driver, "cloud_workbooks", return_value=[]),
            patch.object(driver, "push_notebook", side_effect=OSError("push failed")),
            patch.object(driver, "delete_kernel") as delete,
            patch.object(driver, "user_name", return_value="alice"),
        ):
            with self.assertRaisesRegex(OSError, "push failed"):
                driver.run()
        delete.assert_not_called()
        state = json.loads(driver.state_path().read_text(encoding="utf-8"))
        self.assertEqual(state["phase"], "aborted")
        self.assertEqual(state["failed_phase"], "preparing")
        self.assertIn("push failed", state["error"])

    def test_kernel_delete_failure_records_cleanup_failed_and_possible_run(self) -> None:
        canonical = notebook("TARGET_DATE = None\nprint('one')\n")
        fake_ws = FakeWebSocket(fail_after=0)

        def fake_api(method, path, data=None, **kwargs):
            if method == "POST":
                return {"id": "kernel-1"}
            if method == "DELETE":
                raise OSError("delete unavailable")
            raise AssertionError((method, path, data, kwargs))

        with (
            patch.object(driver, "load_notebook", return_value=canonical),
            patch.object(driver, "start_server"),
            patch.object(driver, "acquire_account_run_lock", return_value=object()),
            patch.object(driver, "release_account_run_lock"),
            patch.object(driver, "kernels", return_value=[]),
            patch.object(driver, "cloud_workbooks", return_value=[]),
            patch.object(driver, "push_notebook"),
            patch.object(driver, "api", side_effect=fake_api),
            patch.object(driver, "_ws_connect", return_value=fake_ws),
            patch.object(driver, "user_name", return_value="alice"),
        ):
            with self.assertRaisesRegex(driver.DriverError, "cleanup failed"):
                driver.run()
        state = json.loads(driver.state_path().read_text(encoding="utf-8"))
        self.assertEqual(state["phase"], "cleanup_failed")
        self.assertEqual(state["failed_phase"], "submitting")
        self.assertTrue(state["possibly_running"])
        self.assertEqual(state["kernel"], "kernel-1")
        self.assertIn("delete unavailable", state["cleanup_error"])

    def test_cleanup_failed_state_blocks_run_even_when_kernel_is_missing(self) -> None:
        driver._save_state(
            {
                "run_id": "previous-run",
                "phase": "cleanup_failed",
                "kernel": "kernel-previous",
                "possibly_running": True,
            }
        )
        canonical = notebook("TARGET_DATE = None\n")
        with (
            patch.object(driver, "load_notebook", return_value=canonical),
            patch.object(driver, "start_server") as start,
            patch.object(
                driver, "acquire_account_run_lock", return_value=object()
            ) as acquire,
            patch.object(driver, "release_account_run_lock") as release,
            patch.object(driver, "kernels", return_value=[]) as kernels,
            patch.object(driver, "push_notebook") as push,
        ):
            with self.assertRaisesRegex(driver.DriverError, "recover"):
                driver.run()
        start.assert_not_called()
        acquire.assert_not_called()
        kernels.assert_not_called()
        push.assert_not_called()
        release.assert_not_called()

    def test_cleanup_failed_race_is_blocked_after_account_lock(self) -> None:
        canonical = notebook("TARGET_DATE = None\n")
        unresolved = {
            "run_id": "concurrent-run",
            "phase": "cleanup_failed",
            "kernel": "kernel-concurrent",
        }
        with (
            patch.object(driver, "load_notebook", return_value=canonical),
            patch.object(driver, "_load_state", side_effect=[None, unresolved]),
            patch.object(driver, "start_server") as start,
            patch.object(
                driver, "acquire_account_run_lock", return_value=object()
            ) as acquire,
            patch.object(driver, "release_account_run_lock") as release,
            patch.object(driver, "kernels") as kernels,
            patch.object(driver, "push_notebook") as push,
        ):
            with self.assertRaisesRegex(driver.DriverError, "kernel-concurrent"):
                driver.run()
        start.assert_called_once()
        acquire.assert_called_once()
        release.assert_called_once()
        kernels.assert_not_called()
        push.assert_not_called()

    def test_recover_resolves_cleanup_failed_state_before_next_run(self) -> None:
        driver._save_state(
            {
                "run_id": "previous-run",
                "phase": "cleanup_failed",
                "kernel": "kernel-previous",
                "possibly_running": True,
                "canonical_restored": False,
                "restore_error": "restore unavailable",
            }
        )
        canonical = notebook("TARGET_DATE = None\n")
        with (
            patch.object(driver, "load_notebook", return_value=canonical),
            patch.object(driver, "start_server"),
            patch.object(driver, "acquire_account_run_lock", return_value=object()),
            patch.object(driver, "release_account_run_lock") as release,
            patch.object(driver, "push_notebook") as push,
            patch.object(driver, "delete_kernel") as delete,
        ):
            driver.recover()
        push.assert_called_once_with(canonical)
        delete.assert_called_once_with("kernel-previous")
        release.assert_called_once()
        state = json.loads(driver.state_path().read_text(encoding="utf-8"))
        self.assertEqual(state["phase"], "aborted")
        self.assertTrue(state["canonical_restored"])
        self.assertTrue(state["kernel_deleted"])
        self.assertFalse(state["possibly_running"])
        self.assertIn("recovered_at_utc", state)

    def test_restore_and_kernel_cleanup_failures_are_both_preserved(self) -> None:
        canonical = notebook("TARGET_DATE = None\nprint('one')\n")
        fake_ws = FakeWebSocket()
        pushes = 0

        def fake_push(value):
            nonlocal pushes
            pushes += 1
            if pushes > 1:
                raise OSError("restore unavailable")

        def fake_api(method, path, data=None, **kwargs):
            if method == "POST":
                return {"id": "kernel-1"}
            if method == "DELETE":
                raise OSError("delete unavailable")
            raise AssertionError((method, path, data, kwargs))

        with (
            patch.object(driver, "load_notebook", return_value=canonical),
            patch.object(driver, "start_server"),
            patch.object(driver, "acquire_account_run_lock", return_value=object()),
            patch.object(driver, "release_account_run_lock"),
            patch.object(driver, "kernels", return_value=[]),
            patch.object(driver, "cloud_workbooks", return_value=[]),
            patch.object(driver, "push_notebook", side_effect=fake_push),
            patch.object(driver, "api", side_effect=fake_api),
            patch.object(driver, "_ws_connect", return_value=fake_ws),
            patch.object(driver, "user_name", return_value="alice"),
        ):
            with self.assertRaisesRegex(
                driver.DriverError,
                "cloud notebook may retain the historical-date override.*kernel cleanup failed",
            ):
                driver.run("2026-08-07")
        state = json.loads(driver.state_path().read_text(encoding="utf-8"))
        self.assertEqual(state["phase"], "cleanup_failed")
        self.assertFalse(state["canonical_restored"])
        self.assertIn("restore unavailable", state["restore_error"])
        self.assertIn("delete unavailable", state["cleanup_error"])

    def test_failed_cloud_restore_blocks_next_run_after_kernel_is_deleted(self) -> None:
        canonical = notebook("TARGET_DATE = None\nprint('one')\n")
        fake_ws = FakeWebSocket()
        pushes = 0
        deletes = []

        def fake_push(value):
            nonlocal pushes
            pushes += 1
            if pushes > 1:
                raise OSError("restore unavailable")

        def fake_api(method, path, data=None, **kwargs):
            if method == "POST":
                return {"id": "kernel-1"}
            if method == "DELETE":
                deletes.append(path)
                return None
            raise AssertionError((method, path, data, kwargs))

        with (
            patch.object(driver, "load_notebook", return_value=canonical),
            patch.object(driver, "start_server"),
            patch.object(driver, "acquire_account_run_lock", return_value=object()),
            patch.object(driver, "release_account_run_lock"),
            patch.object(driver, "kernels", return_value=[]),
            patch.object(driver, "cloud_workbooks", return_value=[]),
            patch.object(driver, "push_notebook", side_effect=fake_push),
            patch.object(driver, "api", side_effect=fake_api),
            patch.object(driver, "_ws_connect", return_value=fake_ws),
            patch.object(driver, "user_name", return_value="alice"),
        ):
            with self.assertRaises(driver.DriverError):
                driver.run("2026-08-07")
        self.assertEqual(pushes, 3)
        self.assertEqual(len(deletes), 1)
        state = json.loads(driver.state_path().read_text(encoding="utf-8"))
        self.assertEqual(state["phase"], "cleanup_failed")
        self.assertFalse(state["possibly_running"])
        self.assertTrue(state["kernel_deleted"])
        self.assertFalse(state["canonical_restored"])
        self.assertIn("restore unavailable", state["restore_error"])

    def test_busy_kernel_blocks_duplicate_run_before_push(self) -> None:
        canonical = notebook("TARGET_DATE = None\n")
        with (
            patch.object(driver, "load_notebook", return_value=canonical),
            patch.object(driver, "start_server"),
            patch.object(driver, "acquire_account_run_lock", return_value=object()),
            patch.object(driver, "release_account_run_lock") as release,
            patch.object(
                driver,
                "kernels",
                return_value=[{"id": "busy-1", "execution_state": "busy"}],
            ),
            patch.object(driver, "push_notebook") as push,
        ):
            with self.assertRaises(driver.DriverError):
                driver.run()
        push.assert_not_called()
        release.assert_called_once()

    def test_state_failure_after_kernel_creation_deletes_kernel(self) -> None:
        canonical = notebook("TARGET_DATE = None\nprint('one')\n")
        deletes = []

        def fake_api(method, path, data=None, **kwargs):
            if method == "POST":
                return {"id": "kernel-1"}
            if method == "DELETE":
                deletes.append(path)
                return None
            raise AssertionError((method, path, data, kwargs))

        with (
            patch.object(driver, "load_notebook", return_value=canonical),
            patch.object(driver, "start_server"),
            patch.object(driver, "acquire_account_run_lock", return_value=object()),
            patch.object(driver, "release_account_run_lock"),
            patch.object(driver, "kernels", return_value=[]),
            patch.object(driver, "cloud_workbooks", return_value=[]),
            patch.object(driver, "push_notebook"),
            patch.object(driver, "api", side_effect=fake_api),
            patch.object(
                driver,
                "_save_state",
                side_effect=[None, OSError("disk full"), None],
            ) as save_state,
            patch.object(driver, "user_name", return_value="alice"),
        ):
            with self.assertRaises(OSError):
                driver.run()
        self.assertEqual(len(deletes), 1)
        self.assertEqual(save_state.call_count, 3)

    def test_freshness_rejects_unchanged_cloud_version(self) -> None:
        item = {
            "name": "supermind_full_20260807.xlsx",
            "last_modified": "2026-08-10T12:00:00Z",
        }
        state = {
            "phase": "submitted",
            "target_date": "2026-08-07",
            "fired_at_utc": "2026-08-10T11:00:00Z",
            "baseline_versions": {item["name"]: "2026-08-10T12:00:00+00:00"},
        }
        with self.assertRaises(driver.DriverError):
            driver._require_fresh_result(item, state)

    def test_freshness_accepts_new_or_updated_cloud_version(self) -> None:
        new_item = {
            "name": "supermind_full_20260807.xlsx",
            "last_modified": "2026-08-10T12:00:01Z",
        }
        state = {
            "phase": "submitted",
            "target_date": "2026-08-07",
            "fired_at_utc": "2026-08-10T12:00:00Z",
            "baseline_versions": {new_item["name"]: "2026-08-09T12:00:00Z"},
        }
        driver._require_fresh_result(new_item, state)
        state["baseline_versions"] = {}
        driver._require_fresh_result(new_item, state)

    def test_freshness_rejects_time_regression(self) -> None:
        item = {
            "name": "supermind_full_20260807.xlsx",
            "last_modified": "2026-08-09T12:00:00Z",
        }
        state = {
            "phase": "submitted",
            "target_date": "2026-08-07",
            "fired_at_utc": "2026-08-08T12:00:00Z",
            "baseline_versions": {item["name"]: "2026-08-10T12:00:00Z"},
        }
        with self.assertRaises(driver.DriverError):
            driver._require_fresh_result(item, state)

    def test_freshness_rejects_new_filename_older_than_run(self) -> None:
        item = {
            "name": "supermind_full_20260807.xlsx",
            "last_modified": "2026-08-09T12:00:00Z",
        }
        state = {
            "phase": "submitted",
            "target_date": "2026-08-07",
            "fired_at_utc": "2026-08-10T12:00:00Z",
            "baseline_versions": {},
        }
        with self.assertRaises(driver.DriverError):
            driver._require_fresh_result(item, state)

    def test_fetch_validation_failure_preserves_existing_good_file(self) -> None:
        day_dir = Path(self.tempdir.name) / "2026-08-07"
        day_dir.mkdir()
        destination = day_dir / "supermind_full_20260807.xlsx"
        destination.write_bytes(b"old-good")
        item = {"name": destination.name, "last_modified": "v2"}
        with (
            patch.object(driver, "start_server"),
            patch.object(driver, "cloud_workbooks", return_value=[item]),
            patch.object(driver, "api", return_value=b"new-bad"),
            patch.object(
                driver,
                "validate_workbook",
                return_value=SimpleNamespace(passed=False, checks=[]),
            ),
            patch.object(driver, "print_result"),
            patch.object(driver, "user_name", return_value="alice"),
        ):
            with self.assertRaises(driver.DriverError):
                driver.fetch("2026-08-07", allow_existing=True)
        self.assertEqual(destination.read_bytes(), b"old-good")
        diagnostics = list(day_dir.glob("*.invalid-*.xlsx"))
        self.assertEqual(len(diagnostics), 1)
        self.assertEqual(diagnostics[0].read_bytes(), b"new-bad")

    def test_fetch_replaces_destination_only_after_validation(self) -> None:
        day_dir = Path(self.tempdir.name) / "2026-08-07"
        day_dir.mkdir()
        destination = day_dir / "supermind_full_20260807.xlsx"
        destination.write_bytes(b"old-good")
        item = {"name": destination.name, "last_modified": "v2"}
        source = Path(self.tempdir.name) / "source.xlsx"
        workbook = Workbook()
        workbook.save(source)

        def validate_real_xlsx(path, **kwargs):
            self.assertEqual(path.suffix, ".xlsx")
            opened = load_workbook(path, read_only=True)
            opened.close()
            return SimpleNamespace(passed=True, checks=[])

        with (
            patch.object(driver, "start_server"),
            patch.object(driver, "cloud_workbooks", return_value=[item]),
            patch.object(driver, "api", return_value=source.read_bytes()),
            patch.object(
                driver,
                "validate_workbook",
                side_effect=validate_real_xlsx,
            ),
            patch.object(driver, "print_result"),
            patch.object(driver, "user_name", return_value="alice"),
        ):
            result = driver.fetch("2026-08-07", allow_existing=True)
        self.assertEqual(result, destination)
        opened = load_workbook(destination, read_only=True)
        opened.close()

    def test_fetch_removes_partial_temp_file_after_write_failure(self) -> None:
        item = {
            "name": "supermind_full_20260807.xlsx",
            "last_modified": "2026-08-10T12:00:01Z",
        }
        original_write = Path.write_bytes

        def fail_after_partial_write(path, data):
            if path.name.endswith(".download.xlsx"):
                with path.open("wb") as handle:
                    handle.write(b"partial")
                raise OSError("disk full")
            return original_write(path, data)

        with (
            patch.object(driver, "start_server"),
            patch.object(driver, "cloud_workbooks", return_value=[item]),
            patch.object(driver, "api", return_value=b"download"),
            patch.object(Path, "write_bytes", new=fail_after_partial_write),
            patch.object(driver, "user_name", return_value="alice"),
        ):
            with self.assertRaises(OSError):
                driver.fetch("2026-08-07", allow_existing=True)
        day_dir = Path(self.tempdir.name) / "2026-08-07"
        self.assertEqual(list(day_dir.glob("*.download.xlsx")), [])

    def test_fetch_removes_temp_file_after_final_replace_failure(self) -> None:
        item = {
            "name": "supermind_full_20260807.xlsx",
            "last_modified": "2026-08-10T12:00:01Z",
        }
        original_replace = Path.replace

        def fail_download_replace(path, target):
            if path.name.endswith(".download.xlsx"):
                raise OSError("replace failed")
            return original_replace(path, target)

        with (
            patch.object(driver, "start_server"),
            patch.object(driver, "cloud_workbooks", return_value=[item]),
            patch.object(driver, "api", return_value=b"download"),
            patch.object(
                driver,
                "validate_workbook",
                return_value=SimpleNamespace(passed=True, checks=[]),
            ),
            patch.object(driver, "print_result"),
            patch.object(Path, "replace", new=fail_download_replace),
            patch.object(driver, "user_name", return_value="alice"),
        ):
            with self.assertRaises(OSError):
                driver.fetch("2026-08-07", allow_existing=True)
        day_dir = Path(self.tempdir.name) / "2026-08-07"
        self.assertEqual(list(day_dir.glob("*.download.xlsx")), [])

    def test_exec_connection_failure_still_deletes_kernel(self) -> None:
        with (
            patch.dict(
                sys.modules,
                {"websocket": SimpleNamespace(WebSocketTimeoutException=TimeoutError)},
            ),
            patch.object(driver, "require_dependencies"),
            patch.object(driver, "start_server"),
            patch.object(driver, "user_name", return_value="alice"),
            patch.object(driver, "api", return_value={"id": "kernel-1"}),
            patch.object(driver, "_ws_connect", side_effect=OSError("handshake failed")),
            patch.object(driver, "delete_kernel") as delete,
        ):
            with self.assertRaises(OSError):
                driver.exec_code("1 + 1")
        delete.assert_called_once_with("kernel-1")

    def test_exec_remote_error_is_nonzero_and_deletes_kernel(self) -> None:
        reply = json.dumps(
            {"msg_type": "execute_reply", "content": {"status": "error"}}
        )
        websocket = FakeWebSocket(received=[reply])
        with (
            patch.dict(
                sys.modules,
                {"websocket": SimpleNamespace(WebSocketTimeoutException=TimeoutError)},
            ),
            patch.object(driver, "require_dependencies"),
            patch.object(driver, "start_server"),
            patch.object(driver, "user_name", return_value="alice"),
            patch.object(driver, "api", return_value={"id": "kernel-1"}),
            patch.object(driver, "_ws_connect", return_value=websocket),
            patch.object(driver, "delete_kernel") as delete,
        ):
            with self.assertRaises(driver.DriverError):
                driver.exec_code("raise RuntimeError()")
        self.assertTrue(websocket.closed)
        delete.assert_called_once_with("kernel-1")

    def test_exec_cleanup_failure_is_reported(self) -> None:
        reply = json.dumps({"msg_type": "execute_reply", "content": {"status": "ok"}})
        websocket = FakeWebSocket(received=[reply])
        with (
            patch.dict(
                sys.modules,
                {"websocket": SimpleNamespace(WebSocketTimeoutException=TimeoutError)},
            ),
            patch.object(driver, "require_dependencies"),
            patch.object(driver, "start_server"),
            patch.object(driver, "user_name", return_value="alice"),
            patch.object(driver, "api", return_value={"id": "kernel-1"}),
            patch.object(driver, "_ws_connect", return_value=websocket),
            patch.object(driver, "delete_kernel", side_effect=OSError("delete failed")),
        ):
            with self.assertRaisesRegex(driver.DriverError, "cleanup failed"):
                driver.exec_code("1 + 1")

    def test_pull_cannot_overwrite_any_file_inside_skill(self) -> None:
        with self.assertRaises(driver.DriverError):
            driver.pull(driver.PACKAGED_NOTEBOOK, force=True)

    def test_pull_cannot_write_to_plugin_sibling(self) -> None:
        with self.assertRaises(driver.DriverError):
            driver.pull(driver.PLUGIN_DIR / "snapshot.ipynb", force=True)

    def test_parser_has_no_arbitrary_notebook_option(self) -> None:
        with redirect_stderr(io.StringIO()), self.assertRaises(SystemExit):
            driver.build_parser().parse_args(["--notebook", "/tmp/a.ipynb", "status"])

    def test_watch_deletes_idle_kernel(self) -> None:
        driver._save_state({"kernel": "kernel-1", "phase": "submitted"})
        websocket = FakeWebSocket(received=[TimeoutError()])
        busy = [{"id": "kernel-1", "execution_state": "busy"}]
        idle = [{"id": "kernel-1", "execution_state": "idle"}]
        with (
            patch.dict(
                sys.modules,
                {"websocket": SimpleNamespace(WebSocketTimeoutException=TimeoutError)},
            ),
            patch.object(driver, "require_dependencies"),
            patch.object(driver, "kernels", side_effect=[busy, idle]),
            patch.object(driver, "_ws_connect", return_value=websocket),
            patch.object(driver, "delete_kernel") as delete,
        ):
            driver.watch(30)
        self.assertTrue(websocket.closed)
        delete.assert_called_once_with("kernel-1")

    def test_watch_close_failure_still_deletes_idle_kernel(self) -> None:
        driver._save_state({"kernel": "kernel-1", "phase": "submitted"})
        websocket = FakeWebSocket(received=[TimeoutError()])
        websocket.close = lambda: (_ for _ in ()).throw(OSError("close failed"))
        busy = [{"id": "kernel-1", "execution_state": "busy"}]
        idle = [{"id": "kernel-1", "execution_state": "idle"}]
        with (
            patch.dict(
                sys.modules,
                {"websocket": SimpleNamespace(WebSocketTimeoutException=TimeoutError)},
            ),
            patch.object(driver, "require_dependencies"),
            patch.object(driver, "kernels", side_effect=[busy, idle]),
            patch.object(driver, "_ws_connect", return_value=websocket),
            patch.object(driver, "delete_kernel") as delete,
        ):
            with self.assertRaisesRegex(driver.DriverError, "close failed"):
                driver.watch(30)
        delete.assert_called_once_with("kernel-1")

    def test_watch_never_attaches_to_untracked_kernel(self) -> None:
        driver._save_state({"kernel": "our-finished-kernel", "phase": "submitted"})
        unrelated = [{"id": "interactive-kernel", "execution_state": "idle"}]
        with (
            patch.object(driver, "kernels", return_value=unrelated),
            patch.object(driver, "_ws_connect") as connect,
            patch.object(driver, "delete_kernel") as delete,
        ):
            driver.watch(30)
        connect.assert_not_called()
        delete.assert_not_called()

    def test_watch_without_runstate_never_attaches_to_active_kernel(self) -> None:
        unrelated = [{"id": "interactive-kernel", "execution_state": "idle"}]
        with (
            patch.object(driver, "kernels", return_value=unrelated),
            patch.object(driver, "_ws_connect") as connect,
            patch.object(driver, "delete_kernel") as delete,
        ):
            with self.assertRaises(driver.DriverError):
                driver.watch(30)
        connect.assert_not_called()
        delete.assert_not_called()

    def test_iopub_output_redacts_cached_token(self) -> None:
        driver._TOKEN_CACHE = "secret-token"
        output = io.StringIO()
        with redirect_stdout(output):
            driver._print_iopub(
                {"msg_type": "stream", "content": {"text": "value=secret-token\n"}}
            )
        self.assertNotIn("secret-token", output.getvalue())
        self.assertIn("<redacted>", output.getvalue())

    def test_account_lock_blocks_second_local_run(self) -> None:
        with patch.object(driver, "user_name", return_value="lock-test-user"):
            first = driver.acquire_account_run_lock()
            try:
                with self.assertRaises(driver.DriverError):
                    driver.acquire_account_run_lock()
            finally:
                driver.release_account_run_lock(first)

    def test_configure_rejects_token_inside_output_tree(self) -> None:
        output = Path(self.tempdir.name) / "output"
        args = SimpleNamespace(
            base_url=driver.DEFAULT_BASE_URL,
            output_dir=output,
            token_file=output / "token.txt",
            initial_watch_seconds=60,
        )
        with self.assertRaises(driver.DriverError):
            driver.configure(args)

    def test_configure_rejects_temporary_run_output_without_opt_in(self) -> None:
        args = SimpleNamespace(
            base_url=driver.DEFAULT_BASE_URL,
            output_dir=Path(self.tempdir.name) / "output",
            token_file=None,
            initial_watch_seconds=60,
            command="run",
            allow_temporary_output=False,
        )
        with self.assertRaisesRegex(driver.DriverError, "temporary output"):
            driver.configure(args)

    def test_configure_rejects_temporary_fetch_output_without_opt_in(self) -> None:
        args = SimpleNamespace(
            base_url=driver.DEFAULT_BASE_URL,
            output_dir=Path(self.tempdir.name) / "output",
            token_file=None,
            command="fetch",
            allow_temporary_output=False,
        )
        with self.assertRaisesRegex(driver.DriverError, "temporary output"):
            driver.configure(args)

    def test_configure_allows_explicit_temporary_run_output(self) -> None:
        output = Path(self.tempdir.name) / "output"
        args = SimpleNamespace(
            base_url=driver.DEFAULT_BASE_URL,
            output_dir=output,
            token_file=None,
            initial_watch_seconds=60,
            command="run",
            allow_temporary_output=True,
        )
        with redirect_stdout(io.StringIO()):
            driver.configure(args)
        self.assertEqual(driver.CONFIG.output_dir, output.resolve())

    def test_configure_allows_explicit_temporary_fetch_output(self) -> None:
        output = Path(self.tempdir.name) / "output"
        args = SimpleNamespace(
            base_url=driver.DEFAULT_BASE_URL,
            output_dir=output,
            token_file=None,
            command="fetch",
            allow_temporary_output=True,
        )
        with redirect_stdout(io.StringIO()):
            driver.configure(args)
        self.assertEqual(driver.CONFIG.output_dir, output.resolve())


if __name__ == "__main__":
    unittest.main()
